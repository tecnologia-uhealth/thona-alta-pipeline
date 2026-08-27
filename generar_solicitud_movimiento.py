"""
Genera el archivo "Solicitud de Movimiento.xlsm/pdf" listo para subir al
portal Thona (hoja "Solicitud" del template oficial), a partir de los datos
del movimiento (ALTA o BAJA) y calcula automáticamente la vigencia
(día 1 del mes actual -> mismo día un año después, según regla documentada).

Requiere LibreOffice instalado en el servidor para la conversión a PDF
(headless): `soffice --headless --convert-to pdf archivo.xlsx --outdir /tmp`

Uso típico:
    from generar_solicitud_movimiento import generar_solicitud_movimiento

    datos = {
        "no_poliza": "68873-00",
        "ramo": "ACCIDENTES PERSONALES",
        "contratante": "U HEALTH INSURTECH",
        "tipo_poliza": "GRUPO o COLECTIVO",
        "tipo_solicitud": "ENDOSO_A",          # ENDOSO_A = alta, ENDOSO_D = baja
        "tipo_movimiento": "ALTA DE ASEGURADO",  # o "BAJA DE ASEGURADO"
        "observaciones": "DAR DE ALTA ASEGURADOS. GRACIAS",
    }
    xlsx_path, pdf_path = generar_solicitud_movimiento(datos, salida_base="/tmp/Solicitud_ORD1234")
"""

import os
import subprocess
from datetime import date
from calendar import monthrange
from openpyxl import load_workbook

TEMPLATE_PATH = "Solicitud_Movimientos_template.xlsx"  # plantilla base (copiar la oficial aquí, sin macros)

# Celdas mapeadas de la plantilla oficial (hoja "Solicitud")
CELL_CONTRATANTE = "B9"
CELL_NO_POLIZA = "B12"
CELL_RAMO = "M12"
CELL_INICIO_VIGENCIA = "V12"
CELL_FIN_VIGENCIA = "AE12"
CELL_TIPO_POLIZA = "B15"
CELL_TIPO_SOLICITUD = "I15"
CELL_TIPO_MOVIMIENTO = "U15"
CELL_OBSERVACIONES = "B52"


def calcular_vigencia(hoy: date = None) -> tuple[date, date]:
    """
    Regla documentada: inicio de vigencia = día 1 del mes actual,
    fin de vigencia = un año después.
    """
    hoy = hoy or date.today()
    inicio = date(hoy.year, hoy.month, 1)
    fin = date(hoy.year + 1, hoy.month, 1)
    # último día del mes de "fin" menos 1 día para que sea "un año después" exacto del día 1
    return inicio, fin


def generar_solicitud_movimiento(datos: dict, salida_base: str, template_path: str = TEMPLATE_PATH,
                                  convertir_pdf: bool = True) -> tuple[str, str | None]:
    """
    datos: dict con no_poliza, ramo, contratante, tipo_poliza, tipo_solicitud,
           tipo_movimiento, observaciones (opcional)
    salida_base: ruta sin extensión, ej. "/tmp/Solicitud_ORD1234" -> genera .xlsx y .pdf
    """
    wb = load_workbook(template_path, keep_vba=False)
    ws = wb["Solicitud"]

    ws[CELL_CONTRATANTE] = datos["contratante"]
    ws[CELL_NO_POLIZA] = datos["no_poliza"]
    ws[CELL_RAMO] = datos["ramo"]

    inicio, fin = calcular_vigencia()
    ws[CELL_INICIO_VIGENCIA] = inicio
    ws[CELL_FIN_VIGENCIA] = fin
    ws[CELL_INICIO_VIGENCIA].number_format = "DD/MM/YYYY"
    ws[CELL_FIN_VIGENCIA].number_format = "DD/MM/YYYY"

    ws[CELL_TIPO_POLIZA] = datos.get("tipo_poliza", "GRUPO o COLECTIVO")
    ws[CELL_TIPO_SOLICITUD] = datos["tipo_solicitud"]      # ENDOSO_A (alta) / ENDOSO_D (baja)
    ws[CELL_TIPO_MOVIMIENTO] = datos["tipo_movimiento"]    # ALTA DE ASEGURADO / BAJA DE ASEGURADO
    if datos.get("observaciones"):
        ws[CELL_OBSERVACIONES] = datos["observaciones"]

    xlsx_path = f"{salida_base}.xlsx"
    wb.save(xlsx_path)

    pdf_path = None
    if convertir_pdf:
        # El workbook tiene 6 hojas (Listas, Endoso A, Endoso B, Endoso D, Endoso 0, Solicitud).
        # Solo la hoja "Solicitud" debe ir al PDF. Ocultar hojas NO es suficiente:
        # LibreOffice --convert-to pdf las exporta de todas formas. Por eso se genera
        # un archivo aparte (solo para el PDF) donde las demás hojas se BORRAN.
        wb_pdf = load_workbook(xlsx_path)
        for hoja in list(wb_pdf.sheetnames):
            if hoja != "Solicitud":
                del wb_pdf[hoja]
        xlsx_solo_solicitud = f"{salida_base}__solo_solicitud.xlsx"
        wb_pdf.save(xlsx_solo_solicitud)

        outdir = "/".join(xlsx_path.split("/")[:-1]) or "."
        subprocess.run(
            ["soffice", "--headless", "--convert-to", "pdf", xlsx_solo_solicitud, "--outdir", outdir],
            check=True, timeout=60
        )
        pdf_path = xlsx_solo_solicitud.rsplit(".", 1)[0] + ".pdf"
        os.remove(xlsx_solo_solicitud)  # era un archivo intermedio, no se necesita después

    return xlsx_path, pdf_path


if __name__ == "__main__":
    ejemplo = {
        "no_poliza": "68873-00",
        "ramo": "ACCIDENTES PERSONALES",
        "contratante": "U HEALTH INSURTECH",
        "tipo_poliza": "GRUPO o COLECTIVO",
        "tipo_solicitud": "ENDOSO_A",
        "tipo_movimiento": "ALTA DE ASEGURADO",
        "observaciones": "DAR DE ALTA ASEGURADOS. GRACIAS",
    }
    xlsx, pdf = generar_solicitud_movimiento(ejemplo, salida_base="/tmp/Solicitud_TEST", convertir_pdf=False)
    print("Generado:", xlsx, pdf)
