"""
Genera el archivo "Layout Asegurados.xlsx" listo para subir al portal Thona,
a partir de los datos de un asegurado nuevo (o una lista de ellos), tomando
como base la plantilla oficial (misma estructura de columnas y encabezado).

Uso típico dentro del pipeline:
    from generar_layout_asegurados import generar_layout_asegurados

    asegurados = [
        {
            "subgrupo": 1,
            "nombres": "MARIA FERNANDA",
            "apellido_paterno": "GOMEZ",
            "apellido_materno": "LOPEZ",
            "fecha_nacimiento": "1990-05-12",   # YYYY-MM-DD
            "genero": "F",                       # "M" o "F"
            "suma_asegurada": None,               # opcional
        }
    ]
    ruta = generar_layout_asegurados(asegurados, salida="/tmp/Layout_Asegurados_ORD1234.xlsx")
"""

from datetime import datetime
from openpyxl import load_workbook

TEMPLATE_PATH = "Layout_Asegurados_template.xlsx"  # plantilla base (copiar la oficial aquí)
HEADER_ROW = 10   # fila de encabezados (Subgrupo, Nombre(s), ...)
FIRST_DATA_ROW = 11


def _parse_fecha(fecha):
    if isinstance(fecha, datetime):
        return fecha
    return datetime.strptime(fecha, "%Y-%m-%d")


def generar_layout_asegurados(asegurados: list[dict], salida: str, template_path: str = TEMPLATE_PATH) -> str:
    """
    asegurados: lista de dicts con las llaves:
        subgrupo, nombres, apellido_paterno, apellido_materno,
        fecha_nacimiento (YYYY-MM-DD o datetime), genero (M/F), suma_asegurada (opcional)
    salida: ruta de archivo .xlsx a generar
    """
    wb = load_workbook(template_path)
    ws = wb.active

    # ⚠️ FIX CRÍTICO: la plantilla base trae filas de ejemplo con personas
    # reales cargadas (José de Jesús Salgado, Alma Delia Antillón, etc.).
    # Sin este borrado, cada alta real se mandaba con esas 10+ personas de
    # más pegadas debajo del asegurado nuevo. Se limpia todo el rango de
    # datos antes de escribir, para que el Excel final solo contenga a
    # quien realmente se está dando de alta en esta solicitud.
    max_fila_con_datos = ws.max_row
    for fila in range(FIRST_DATA_ROW, max_fila_con_datos + 1):
        for columna in range(1, 8):  # A-G: Subgrupo..Sueldo Mensual
            ws.cell(row=fila, column=columna).value = None

    row = FIRST_DATA_ROW
    for a in asegurados:
        ws.cell(row=row, column=1, value=a.get("subgrupo", 1))
        ws.cell(row=row, column=2, value=a["nombres"])
        ws.cell(row=row, column=3, value=a["apellido_paterno"])
        ws.cell(row=row, column=4, value=a["apellido_materno"])
        ws.cell(row=row, column=5, value=_parse_fecha(a["fecha_nacimiento"]))
        ws.cell(row=row, column=5).number_format = "DD/MM/YYYY"
        ws.cell(row=row, column=6, value=a["genero"])
        if a.get("suma_asegurada") is not None:
            ws.cell(row=row, column=7, value=a["suma_asegurada"])
        row += 1

    wb.save(salida)
    return salida


if __name__ == "__main__":
    # Prueba rápida con datos de ejemplo
    ejemplo = [{
        "subgrupo": 1,
        "nombres": "JUAN CARLOS",
        "apellido_paterno": "PEREZ",
        "apellido_materno": "TORRES",
        "fecha_nacimiento": "1995-03-21",
        "genero": "M",
    }]
    ruta = generar_layout_asegurados(ejemplo, salida="/tmp/Layout_Asegurados_TEST.xlsx")
    print("Generado:", ruta)
